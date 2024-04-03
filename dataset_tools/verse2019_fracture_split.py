'''
Description: this file arrording the spine fracture xlsx to split verse2019.The xlsx file can be get in https://www.ncbi.nlm.nih.gov/pmc/articles/PMC8082364/
version: 1.0
Author: Shuai Lei
Date: 2024-03-23 11:58:45
LastEditors: ShuaiLei
LastEditTime: 2024-04-03 03:03:55
'''
import sys
import os
current_file_path = os.path.abspath(__file__)
project_root = os.path.dirname(os.path.dirname(current_file_path))
if project_root not in sys.path:
    sys.path.insert(0, os.path.dirname(sys.path[0]))
from openpyxl import load_workbook
from io_tools.file_management import load_json_file, save_json_file, create_folder, join
from pycocotools.coco import COCO
import shutil


def split_fracture_ct_dataset(xlsx_file, choosed_vertebrae_label_list):
    """
    load xlsx file to split verse2019 dataset.
    param: xlsx_file: A document that records specific fracture information.
    param: choosed_vertebrae_label_list: The list of selected vertebrae have fracture vertebrae.
    """
    fracture_grad1_dataset = []
    fracture_grad2_dataset = []
    fracture_grad3_dataset = []
    normal_dataset = []
    choosed_vertebrae_label_fx_g_list = [label + "_fx-g" for label in choosed_vertebrae_label_list]
    workbbok = load_workbook(xlsx_file, data_only=True)
    sheet = workbbok.active
    # get header
    header = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # get ct name
        if row[header.index('verse_ID')] != None and row[header.index('verse_ID')] != row[header.index('subject_ID')]:
            ct_name = "sub-verse{:03d}_split-verse{:03d}_ct.nii.gz".format(row[header.index('subject_ID')], row[header.index('verse_ID')])
        if row[header.index('verse_ID')] != None and row[header.index('verse_ID')] == row[header.index('subject_ID')]:  
            ct_name = "sub-verse{:03d}_ct.nii.gz".format(row[header.index('verse_ID')])
        if row[0] != None and row[header.index('N_Fx')] != 0:
            fx_g_id_list = [row[header.index(col)] for col in choosed_vertebrae_label_fx_g_list]
            if 3 in fx_g_id_list:
                fracture_grad3_dataset.append(ct_name)
            if 2 in fx_g_id_list and 3 not in fx_g_id_list:
                fracture_grad2_dataset.append(ct_name)
            if 2 not in fx_g_id_list and 3 not in fx_g_id_list:
                fracture_grad1_dataset.append(ct_name)
        if row[0] != None and row[header.index('N_Fx')] == 0:
            normal_dataset.append(ct_name)
    print("CT Facture grad = 3 number is: ", len(fracture_grad3_dataset))
    print(fracture_grad3_dataset)
    print("CT Facture grad = 2 number is: ", len(fracture_grad2_dataset))
    print(fracture_grad2_dataset)
    print("CT Facture grad = 1 number is: ", len(fracture_grad1_dataset))
    print(fracture_grad1_dataset)
    print("normal number is: ", len(normal_dataset))
    print(normal_dataset)


def fracture_xlsx_file_to_json(xlsx_file, choosed_vertebrae_label_list, fracture_annotation_file):
    """
    conver the T9-L6 to fracture and normal label.
    param: xlsx_file: The verse2019 fracture record file.
    param: choosed_vertebrae_label_list: The drr images detection annotation file.
    param: fracture_annotation_file: The fracture and normal label annotation file.
    """
    fracture_grad3_info_dict = {}
    normal_info_dict = {}
    choosed_vertebrae_label_fx_g_list = [label + "_fx-g" for label in choosed_vertebrae_label_list]
    no_choosed_vertebrae_label_fx_g_list = [label + "_fx-g" for label in ["T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8"]]
    workbbok = load_workbook(xlsx_file, data_only=True)
    sheet = workbbok.active
    # get header
    header = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # get ct name
        if row[header.index('verse_ID')] != None:
            ct_name = "sub-verse{:03d}".format(row[header.index('verse_ID')])
            choosed_categories_fracture_dict = {col.split("_")[0]: row[header.index(col)] for col in choosed_vertebrae_label_fx_g_list}
            no_choosed_categories_fracture_dict = {col.split("_")[0]: row[header.index(col)] for col in no_choosed_vertebrae_label_fx_g_list}
            
            has_T1_T8 = False
            for value in no_choosed_categories_fracture_dict.values():
                if value != None:
                    has_T1_T8 = True
            if row[header.index('N_Fx')] != 0 and 3 in choosed_categories_fracture_dict.values():
                if has_T1_T8:
                    fracture_grad3_info_dict[ct_name + "bottom.nii.gz"] = choosed_categories_fracture_dict
                else:
                    fracture_grad3_info_dict[ct_name + ".nii.gz"] = choosed_categories_fracture_dict
            if row[header.index('N_Fx')] == 0:
                if has_T1_T8:
                    normal_info_dict[ct_name + "bottom.nii.gz"] = choosed_categories_fracture_dict
                else:
                    normal_info_dict[ct_name + ".nii.gz"] = choosed_categories_fracture_dict
    
    for ct_name, fracture_dict in fracture_grad3_info_dict.items():
        for label, fracture_id in fracture_dict.items():
            if fracture_id == 3:
                fracture_dict[label] = "fracture"
            if fracture_id == 0 or fracture_id == 1 or fracture_id == 2:
                fracture_dict[label] = "normal"

    print(fracture_grad3_info_dict)
    save_json_file(fracture_grad3_info_dict, fracture_annotation_file)



def conver_gt_bbox_to_fracture_annotation_file(annotation_file, fracture_record_json_file, fracture_annotation_file):
    """
    The function will be used to conver annotation file which categories is  ["T9", "T10", "T11", "T12","L1", "L2", "L3", "L4", "L5", "L6"] to 
    fracture annotation file while categories is ["normal" , "fracture"]
    param: annotation_file: the ["T9", "T10", "T11", "T12","L1", "L2", "L3", "L4", "L5", "L6"] json file.
    param: fracture_record_json_file: the json file which record fracture grade info.
    param: fracture_annotation_file: the converd ["normal" , "fracture"] json file.
    """
    categories = [{"id": 1,
                    "name": "normal",
                    "supercategory": "vertebrae"
                    },
                    {
                    "id": 2,
                    "name": "fracture",
                    "supercategory": "vertebrae"
                    }]
    
    gt = COCO(annotation_file)
    fracture_data = load_json_file(fracture_record_json_file)
    for image in gt.dataset["images"]:
        single_fracture_data = fracture_data[image["ct_name"]]
        for ann in gt.imgToAnns[image["id"]]:
            ann["category_name"] = single_fracture_data[ann["category_name"]]
            ann["category_id"] = 1 if ann["category_name"] == "normal" else 2
    gt.dataset["categories"] = categories
    save_json_file(gt.dataset, fracture_annotation_file)


def split_cut_images2_fracture_and_normal_dataset(fracture_record_json_file,cut_images_folder, fracture_folder, normal_folder):
    """
    The function will be used to generate fracture and normal dataset according the fracture json file.
    param: fracture_record_json_file: The json file which according paper xlsx generate.
    param: cut_images_folder: The cut images folder path.
    param: fracture_folder: The fracture images save folder path.
    param: normal_folder: The normal images save folder path.
    """
    create_folder(fracture_folder)
    create_folder(normal_folder)
    ct_fracture_data = load_json_file(fracture_record_json_file)
    for file_name in os.listdir(cut_images_folder):
        ct_name = file_name.split("_")[0] + ".nii.gz"
        vertebrae_label = file_name.split("_")[2]
        if ct_fracture_data[ct_name][vertebrae_label] == "normal":
            shutil.copy(join(cut_images_folder, file_name), join(normal_folder, file_name))
        if ct_fracture_data[ct_name][vertebrae_label] == "fracture":
            shutil.copy(join(cut_images_folder, file_name), join(fracture_folder, file_name))
    print("split dataset to fracture and normal folder successfully!")


if __name__ == "__main__":
    # split_fracture_ct_dataset("data/verse2019/verse2019_fracture_grading_info.xlsx", 
    #                           ["T9", "T10", "T11", "T12","L1", "L2", "L3", "L4", "L5", "L6"])

    fracture_xlsx_file_to_json("data/verse2019/verse2019_fracture_grading_info.xlsx",
                              ["T9", "T10", "T11", "T12","L1", "L2", "L3", "L4", "L5", "L6"],
                               "data/verse2019_segmentation_dataset/LA/verse2019_fracture_grading_info.json")
    conver_gt_bbox_to_fracture_annotation_file("data/verse2019_segmentation_dataset/LA/gt_bbox.json", 
                                               "data/verse2019_segmentation_dataset/LA/verse2019_fracture_grading_info.json",
                                               "data/verse2019_segmentation_dataset/LA/fracture_gt_bbox.json")
    split_cut_images2_fracture_and_normal_dataset("data/verse2019_segmentation_dataset/LA/verse2019_fracture_grading_info.json",
                                                  "data/verse2019_segmentation_dataset/LA/cut_images",
                                                  "data/verse2019_segmentation_dataset/LA/fracture_images",
                                                  "data/verse2019_segmentation_dataset/LA/normal_images")
    print("normal images number = ", len(os.listdir("data/verse2019_segmentation_dataset/LA/normal_images")))
    print("fracture images number = ", len(os.listdir("data/verse2019_segmentation_dataset/LA/fracture_images")))